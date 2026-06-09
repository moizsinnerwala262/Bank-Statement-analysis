"""Excel report builder - produces 6-sheet BSA report."""
from calendar import monthrange
from datetime import date, timedelta
from io import BytesIO
from typing import List, Tuple
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

from parsers.base import ParsedStatement
from analyzers.abb import ABBResult
from analyzers.emi import DetectedEMI
from analyzers.party import PartyAggregate


FONT = "Arial"
HEADER_FILL = PatternFill("solid", start_color="1F4E78")
SUB_FILL = PatternFill("solid", start_color="D9E1F2")
CARRY_FILL = PatternFill("solid", start_color="FFF2CC")
ABB_FILL = PatternFill("solid", start_color="C6EFCE")
EMI_FILL = PatternFill("solid", start_color="FCE4D6")
FLAG_FILL = PatternFill("solid", start_color="F4B084")
CREDIT_FILL = PatternFill("solid", start_color="E2EFDA")
DEBIT_FILL = PatternFill("solid", start_color="FCE4D6")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=11)
BOLD = Font(name=FONT, bold=True, size=10)
NORMAL = Font(name=FONT, size=10)
SMALL = Font(name=FONT, size=9)
TITLE_FONT = Font(name=FONT, bold=True, size=14, color="1F4E78")
SUBTITLE = Font(name=FONT, bold=True, size=11, color="1F4E78")
thin = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
MONTH_NAMES = ["", "Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


def build_report(
    stmt: ParsedStatement,
    abb: ABBResult,
    emis: List[DetectedEMI],
    parties: Tuple[List[PartyAggregate], List[PartyAggregate]] = None,
    monthly_stats=None,
    returns_data=None,
) -> bytes:
    wb = Workbook()
    _build_matrix_sheet(wb, abb, monthly_stats)
    _build_statement_sheet(wb, stmt)
    _build_emi_sheet(wb, emis, abb.months)
    if parties is not None:
        _build_party_sheet(wb, parties[0], parties[1], abb.months)
    if monthly_stats:
        _build_monthly_summary_sheet(wb, monthly_stats)
    if returns_data is not None:
        _build_returns_sheet(wb, returns_data)
    _build_summary_sheet(wb, stmt, abb, emis)
    _build_chart_sheet(wb, abb)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_matrix_sheet(wb, abb: ABBResult, monthly_stats=None):
    ws = wb.active
    ws.title = "Daily Balance Matrix"
    months = abb.months

    ws["A1"] = "DAY-WISE EOD BALANCE MATRIX"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(months)+1)
    ws["A2"] = "Yellow cells = carry-forward (no transaction on that day)"
    ws["A2"].font = Font(name=FONT, italic=True, size=9, color="595959")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(months)+1)

    ws.cell(row=4, column=1, value="Day").font = HEADER_FONT
    ws.cell(row=4, column=1).fill = HEADER_FILL
    ws.cell(row=4, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=4, column=1).border = BORDER
    for col_idx, (yr, mo) in enumerate(months, start=2):
        c = ws.cell(row=4, column=col_idx, value=f"{MONTH_NAMES[mo]}-{str(yr)[-2:]}")
        c.font = HEADER_FONT; c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center"); c.border = BORDER

    for day in range(1, 32):
        row = 4 + day
        dc = ws.cell(row=row, column=1, value=day)
        dc.font = BOLD; dc.alignment = Alignment(horizontal="center")
        dc.border = BORDER; dc.fill = SUB_FILL
        for col_idx, (yr, mo) in enumerate(months, start=2):
            days_in_mo = monthrange(yr, mo)[1]
            if day > days_in_mo:
                ws.cell(row=row, column=col_idx).border = BORDER
                continue
            d = date(yr, mo, day)
            if d < abb.period_start or d > abb.period_end:
                ws.cell(row=row, column=col_idx).border = BORDER
                continue
            bal = abb.daily_eod.get(d)
            c = ws.cell(row=row, column=col_idx, value=bal)
            c.number_format = '#,##0.00'
            c.font = NORMAL; c.border = BORDER
            c.alignment = Alignment(horizontal="right")
            if abb.is_carry_fwd.get(d):
                c.fill = CARRY_FILL

    def add_summary(row, label, fill, formula_fn, num_fmt='#,##0.00'):
        lc = ws.cell(row=row, column=1, value=label)
        lc.font = Font(name=FONT, bold=True, size=10); lc.fill = fill
        lc.border = BORDER; lc.alignment = Alignment(horizontal="center")
        for col_idx, (yr, mo) in enumerate(months, start=2):
            col_letter = get_column_letter(col_idx)
            days_in_mo = monthrange(yr, mo)[1]
            c = ws.cell(row=row, column=col_idx, value=formula_fn(col_letter, 5, 4 + days_in_mo))
            c.number_format = num_fmt; c.font = BOLD; c.fill = fill
            c.border = BORDER; c.alignment = Alignment(horizontal="right")

    add_summary(36, "Days in Month", PatternFill("solid", start_color="BDD7EE"),
                lambda col, s, e: f'=COUNT({col}{s}:{col}{e})', '#,##0')
    add_summary(37, "Sum of EOD", PatternFill("solid", start_color="BDD7EE"),
                lambda col, s, e: f'=SUM({col}{s}:{col}{e})')
    add_summary(38, "ABB (Average)", ABB_FILL,
                lambda col, s, e: f'=AVERAGE({col}{s}:{col}{e})')
    add_summary(39, "Min EOD", PatternFill("solid", start_color="FCE4D6"),
                lambda col, s, e: f'=MIN({col}{s}:{col}{e})')
    add_summary(40, "Max EOD", PatternFill("solid", start_color="FCE4D6"),
                lambda col, s, e: f'=MAX({col}{s}:{col}{e})')
    add_summary(41, "Days < ₹1,000", PatternFill("solid", start_color="FFD966"),
                lambda col, s, e: f'=COUNTIFS({col}{s}:{col}{e},"<1000")', '#,##0')

    # New rows 43-47 - monthly Cr/Dr metrics from monthly_stats
    if monthly_stats:
        stats_by_key = {(m.year, m.month): m for m in monthly_stats}

        def write_metric_row(row, label, fill, value_fn, num_fmt='#,##0.00'):
            lc = ws.cell(row=row, column=1, value=label)
            lc.font = Font(name=FONT, bold=True, size=10); lc.fill = fill
            lc.border = BORDER; lc.alignment = Alignment(horizontal="center")
            for col_idx, (yr, mo) in enumerate(months, start=2):
                m = stats_by_key.get((yr, mo))
                val = value_fn(m) if m else 0
                c = ws.cell(row=row, column=col_idx, value=val)
                c.number_format = num_fmt; c.font = BOLD; c.fill = fill
                c.border = BORDER; c.alignment = Alignment(horizontal="right")

        # Spacer row at 42
        ws.cell(row=42, column=1).fill = PatternFill("solid", start_color="FFFFFF")

        CR_FILL = PatternFill("solid", start_color="C6EFCE")  # green-ish
        DR_FILL = PatternFill("solid", start_color="FFC7CE")  # red-ish
        CHARGE_FILL = PatternFill("solid", start_color="FFEB9C")  # yellow

        write_metric_row(43, "Total Credit (₹)", CR_FILL,
                         lambda m: m.cr_amount)
        write_metric_row(44, "# Credit Txns",   CR_FILL,
                         lambda m: m.cr_count, '#,##0')
        write_metric_row(45, "Total Debit (₹)", DR_FILL,
                         lambda m: m.dr_amount)
        write_metric_row(46, "# Debit Txns",    DR_FILL,
                         lambda m: m.dr_count, '#,##0')
        write_metric_row(47, "Bank Charges (₹)", CHARGE_FILL,
                         lambda m: m.bank_charges_amount)
        write_metric_row(48, "# Returns",       CHARGE_FILL,
                         lambda m: m.inward_returns_count + m.outward_returns_count + m.ecs_nach_returns_count,
                         '#,##0')

    last_col_letter = get_column_letter(len(months) + 1)
    ws.conditional_formatting.add(
        f"B5:{last_col_letter}35",
        ColorScaleRule(start_type="num", start_value=0, start_color="F8696B",
                       mid_type="percentile", mid_value=50, mid_color="FFEB84",
                       end_type="max", end_color="63BE7B")
    )
    ws.column_dimensions["A"].width = 6
    for i in range(2, len(months) + 2):
        ws.column_dimensions[get_column_letter(i)].width = 14
    ws.freeze_panes = "B5"


def _build_statement_sheet(wb, stmt: ParsedStatement):
    ws = wb.create_sheet("Statement")
    M = stmt.metadata
    AXIS_RED = PatternFill("solid", start_color="97144D")

    ws["A1"] = M.bank.upper()
    ws["A1"].font = Font(name=FONT, bold=True, size=18, color="FFFFFF")
    ws["A1"].fill = AXIS_RED
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 32
    ws.merge_cells("A1:G1")

    ws["A2"] = "STATEMENT OF ACCOUNT"
    ws["A2"].font = Font(name=FONT, bold=True, size=12, color="FFFFFF")
    ws["A2"].fill = AXIS_RED
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A2:G2")

    holder_block = [
        ("Account Holder", M.account_holder),
        ("Address", M.address),
        ("PAN", M.pan),
        ("Mobile", M.mobile),
        ("Email", M.email),
    ]
    right_block = [
        ("Account No.", M.account_no),
        ("Customer ID", M.customer_id),
        ("IFSC Code", M.ifsc),
        ("MICR Code", M.micr),
        ("Scheme", M.account_type),
    ]
    row = 4
    for i in range(max(len(holder_block), len(right_block))):
        r = row + i
        if i < len(holder_block):
            l, v = holder_block[i]
            ws.cell(row=r, column=1, value=l).font = BOLD
            ws.cell(row=r, column=1).fill = SUB_FILL
            ws.cell(row=r, column=1).border = BORDER
            vc = ws.cell(row=r, column=2, value=v); vc.font = NORMAL
            vc.border = BORDER
            vc.alignment = Alignment(wrap_text=True, vertical="center")
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        if i < len(right_block):
            l, v = right_block[i]
            ws.cell(row=r, column=4, value=l).font = BOLD
            ws.cell(row=r, column=4).fill = SUB_FILL
            ws.cell(row=r, column=4).border = BORDER
            vc = ws.cell(row=r, column=5, value=v); vc.font = NORMAL
            vc.border = BORDER
            ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=7)

    row += max(len(holder_block), len(right_block)) + 1
    period_str = ""
    if M.period_from and M.period_to:
        period_str = f"Statement Period: {M.period_from.strftime('%d-%b-%Y')} to {M.period_to.strftime('%d-%b-%Y')}"
    pc = ws.cell(row=row, column=1, value=period_str)
    pc.font = Font(name=FONT, bold=True, size=11, color="FFFFFF")
    pc.fill = HEADER_FILL
    pc.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    ws.row_dimensions[row].height = 22

    row += 2
    txn_header_row = row
    for col, h in enumerate(["Tran Date", "Chq No", "Particulars", "Debit (₹)", "Credit (₹)", "Balance (₹)", "Branch"], 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = HEADER_FONT; c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center"); c.border = BORDER

    row += 1
    if M.opening_balance is not None:
        c = ws.cell(row=row, column=3, value="OPENING BALANCE")
        c.font = BOLD; c.fill = PatternFill("solid", start_color="E7E6E6")
        c.alignment = Alignment(horizontal="left"); c.border = BORDER
        bc = ws.cell(row=row, column=6, value=M.opening_balance)
        bc.font = BOLD; bc.fill = PatternFill("solid", start_color="E7E6E6")
        bc.number_format = '#,##0.00'; bc.border = BORDER
        bc.alignment = Alignment(horizontal="right")
        for col in [1, 2, 4, 5, 7]:
            ws.cell(row=row, column=col).fill = PatternFill("solid", start_color="E7E6E6")
            ws.cell(row=row, column=col).border = BORDER
        row += 1

    data_start = row
    for t in stmt.transactions:
        dc = ws.cell(row=row, column=1, value=t.date)
        dc.number_format = "dd-mm-yyyy"; dc.font = SMALL; dc.border = BORDER
        dc.alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=2, value=t.cheque_no or "").border = BORDER
        ws.cell(row=row, column=2).font = SMALL
        pc = ws.cell(row=row, column=3, value=t.particulars); pc.font = SMALL; pc.border = BORDER
        if t.debit is not None:
            c = ws.cell(row=row, column=4, value=t.debit)
            c.font = Font(name=FONT, color="C00000", size=9)
            c.number_format = '#,##0.00'; c.alignment = Alignment(horizontal="right")
        else:
            c = ws.cell(row=row, column=4)
        c.border = BORDER
        if t.credit is not None:
            c = ws.cell(row=row, column=5, value=t.credit)
            c.font = Font(name=FONT, color="007F0E", size=9)
            c.number_format = '#,##0.00'; c.alignment = Alignment(horizontal="right")
        else:
            c = ws.cell(row=row, column=5)
        c.border = BORDER
        bc = ws.cell(row=row, column=6, value=t.balance)
        bc.font = SMALL; bc.number_format = '#,##0.00'; bc.border = BORDER
        bc.alignment = Alignment(horizontal="right")
        ic = ws.cell(row=row, column=7, value=t.branch or "")
        ic.font = SMALL; ic.border = BORDER; ic.alignment = Alignment(horizontal="center")
        row += 1
    data_end = row - 1

    tt = ws.cell(row=row, column=3, value="TRANSACTION TOTAL")
    tt.font = BOLD; tt.fill = PatternFill("solid", start_color="FFE699")
    tt.alignment = Alignment(horizontal="left"); tt.border = BORDER
    td = ws.cell(row=row, column=4, value=f"=SUM(D{data_start}:D{data_end})")
    td.font = BOLD; td.fill = PatternFill("solid", start_color="FFE699")
    td.number_format = '#,##0.00'; td.alignment = Alignment(horizontal="right"); td.border = BORDER
    tc = ws.cell(row=row, column=5, value=f"=SUM(E{data_start}:E{data_end})")
    tc.font = BOLD; tc.fill = PatternFill("solid", start_color="FFE699")
    tc.number_format = '#,##0.00'; tc.alignment = Alignment(horizontal="right"); tc.border = BORDER
    for col in [1, 2, 6, 7]:
        ws.cell(row=row, column=col).fill = PatternFill("solid", start_color="FFE699")
        ws.cell(row=row, column=col).border = BORDER
    row += 1

    if M.closing_balance is not None:
        cb = ws.cell(row=row, column=3, value="CLOSING BALANCE")
        cb.font = BOLD; cb.fill = PatternFill("solid", start_color="C6EFCE")
        cb.alignment = Alignment(horizontal="left"); cb.border = BORDER
        cbv = ws.cell(row=row, column=6, value=M.closing_balance)
        cbv.font = BOLD; cbv.fill = PatternFill("solid", start_color="C6EFCE")
        cbv.number_format = '#,##0.00'; cbv.alignment = Alignment(horizontal="right"); cbv.border = BORDER
        for col in [1, 2, 4, 5, 7]:
            ws.cell(row=row, column=col).fill = PatternFill("solid", start_color="C6EFCE")
            ws.cell(row=row, column=col).border = BORDER

    ws.column_dimensions["A"].width = 13
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 10
    ws.freeze_panes = f"A{txn_header_row+1}"


def _build_emi_sheet(wb, emis: List[DetectedEMI], months):
    ws = wb.create_sheet("EMI Extract")

    ws["A1"] = "EMI / RECURRING OBLIGATIONS EXTRACT"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(months)+2)
    ws["A2"] = ("Detection: ACH-DR / ECS / NACH / Loan Recovery debits matched against lender keyword list, validated by 2+ month recurrence. "
                "Loan principal estimates below assume fixed-rate amortization; tenure and rate are inferred from loan-type heuristics.")
    ws["A2"].font = Font(name=FONT, italic=True, size=9, color="595959")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(months)+2)

    row = 4
    ws.cell(row=row, column=1, value="Section 1: EMI Summary + Approximate Loan Estimation").font = SUBTITLE
    row += 1

    headers = ["#", "Lender", "Mode", "Mandate Key", "Typical Day", "Day Var", "Months", "First EMI", "Last EMI",
               "Avg EMI (₹)", "Min (₹)", "Max (₹)", "Var %", "Total Paid (₹)", "Flag",
               "Loan Type", "Confidence", "Approx Principal Low (₹)", "Approx Principal Mid (₹)", "Approx Principal High (₹)"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = HEADER_FONT; c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", wrap_text=True); c.border = BORDER
    ws.row_dimensions[row].height = 30
    row += 1
    emi_start_row = row

    if not emis:
        ws.cell(row=row, column=1, value="No EMI obligations detected in this statement.").font = Font(name=FONT, italic=True, size=10, color="595959")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=20)
        row += 1
    else:
        from analyzers.loan_estimate import estimate_loan
        CONFIDENCE_FILLS = {
            "High": PatternFill("solid", start_color="C6EFCE"),
            "Medium": PatternFill("solid", start_color="FFEB9C"),
            "Low": PatternFill("solid", start_color="FFC7CE"),
        }
        for idx, e in enumerate(emis, 1):
            if e.amt_variance_pct < 5:
                flag_fill = PatternFill("solid", start_color="C6EFCE")
            elif e.amt_variance_pct < 25:
                flag_fill = PatternFill("solid", start_color="FFEB9C")
            else:
                flag_fill = FLAG_FILL

            ws.cell(row=row, column=1, value=idx).alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=2, value=e.lender)
            ws.cell(row=row, column=3, value=e.mode).alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=4, value=(e.norm_mandate or "(no mandate)")[:30])
            ws.cell(row=row, column=5, value=e.typical_day).alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=6, value=e.day_variance).alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=7, value=e.months_seen).alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=8, value=e.first_date).number_format = "dd-mmm-yy"
            ws.cell(row=row, column=9, value=e.last_date).number_format = "dd-mmm-yy"
            ws.cell(row=row, column=10, value=e.avg_emi).number_format = '#,##0.00'
            ws.cell(row=row, column=11, value=e.min_emi).number_format = '#,##0.00'
            ws.cell(row=row, column=12, value=e.max_emi).number_format = '#,##0.00'
            ws.cell(row=row, column=13, value=e.amt_variance_pct / 100).number_format = '0.0%'
            ws.cell(row=row, column=14, value=e.total_paid).number_format = '#,##0.00'
            fc = ws.cell(row=row, column=15, value=e.flag)
            fc.fill = flag_fill; fc.font = Font(name=FONT, size=9)

            # NEW: Loan estimation columns
            est = estimate_loan(e)
            ws.cell(row=row, column=16, value=est.loan_type).font = Font(name=FONT, size=9)
            conf_cell = ws.cell(row=row, column=17, value=est.confidence)
            conf_cell.alignment = Alignment(horizontal="center")
            if est.confidence in CONFIDENCE_FILLS:
                conf_cell.fill = CONFIDENCE_FILLS[est.confidence]
            ws.cell(row=row, column=18, value=est.principal_low).number_format = '#,##0'
            mid_cell = ws.cell(row=row, column=19, value=est.principal_mid)
            mid_cell.number_format = '#,##0'; mid_cell.font = BOLD
            ws.cell(row=row, column=20, value=est.principal_high).number_format = '#,##0'

            for col in range(1, 21):
                cell = ws.cell(row=row, column=col)
                cell.border = BORDER
                if cell.font.size is None:
                    cell.font = NORMAL
            row += 1

        # Total row
        ws.cell(row=row, column=2, value="TOTAL EMI OBLIGATION + ESTIMATED DEBT EXPOSURE").font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
        ws.cell(row=row, column=2).fill = HEADER_FILL
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=9)
        tc = ws.cell(row=row, column=10, value=f"=SUM(J{emi_start_row}:J{emi_start_row + len(emis) - 1})")
        tc.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
        tc.fill = HEADER_FILL; tc.number_format = '"₹"#,##0.00'
        tc.alignment = Alignment(horizontal="right")
        for col in [1, 11, 12, 13, 15, 16, 17]:
            ws.cell(row=row, column=col).fill = HEADER_FILL
        totp = ws.cell(row=row, column=14, value=f"=SUM(N{emi_start_row}:N{emi_start_row + len(emis) - 1})")
        totp.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
        totp.fill = HEADER_FILL; totp.number_format = '"₹"#,##0.00'
        totp.alignment = Alignment(horizontal="right")

        # Sum the principal columns (R/S/T = 18/19/20)
        for col, letter in [(18, "R"), (19, "S"), (20, "T")]:
            c = ws.cell(row=row, column=col,
                        value=f"=SUM({letter}{emi_start_row}:{letter}{emi_start_row + len(emis) - 1})")
            c.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
            c.fill = HEADER_FILL; c.number_format = '"₹"#,##0'
            c.alignment = Alignment(horizontal="right")
        for col in range(1, 21):
            ws.cell(row=row, column=col).border = BORDER
        row += 3

        # Section 2: Date Grid
        ws.cell(row=row, column=1, value="Section 2: EMI Date Grid (DD of debit per month)").font = SUBTITLE
        row += 1
        ws.cell(row=row, column=1, value="#").font = HEADER_FONT
        ws.cell(row=row, column=1).fill = HEADER_FILL
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center"); ws.cell(row=row, column=1).border = BORDER
        ws.cell(row=row, column=2, value="Lender").font = HEADER_FONT
        ws.cell(row=row, column=2).fill = HEADER_FILL
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="center"); ws.cell(row=row, column=2).border = BORDER
        for col_idx, (yr, mo) in enumerate(months, start=3):
            c = ws.cell(row=row, column=col_idx, value=f"{MONTH_NAMES[mo]}-{str(yr)[-2:]}")
            c.font = HEADER_FONT; c.fill = HEADER_FILL
            c.alignment = Alignment(horizontal="center"); c.border = BORDER
        row += 1

        for idx, e in enumerate(emis, 1):
            ws.cell(row=row, column=1, value=idx).alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=1).border = BORDER
            ws.cell(row=row, column=2, value=e.lender).font = NORMAL
            ws.cell(row=row, column=2).border = BORDER
            by_ym = {(it["date"].year, it["date"].month): it["date"].day for it in e.items}
            for col_idx, (yr, mo) in enumerate(months, start=3):
                c = ws.cell(row=row, column=col_idx)
                if (yr, mo) in by_ym:
                    c.value = by_ym[(yr, mo)]
                    c.alignment = Alignment(horizontal="center")
                    c.font = BOLD; c.fill = EMI_FILL
                c.border = BORDER
            row += 1
        row += 2

        # Section 3: Amount Grid
        ws.cell(row=row, column=1, value="Section 3: EMI Amount Grid (₹ debited per month)").font = SUBTITLE
        row += 1
        ws.cell(row=row, column=1, value="#").font = HEADER_FONT
        ws.cell(row=row, column=1).fill = HEADER_FILL
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center"); ws.cell(row=row, column=1).border = BORDER
        ws.cell(row=row, column=2, value="Lender").font = HEADER_FONT
        ws.cell(row=row, column=2).fill = HEADER_FILL
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="center"); ws.cell(row=row, column=2).border = BORDER
        for col_idx, (yr, mo) in enumerate(months, start=3):
            c = ws.cell(row=row, column=col_idx, value=f"{MONTH_NAMES[mo]}-{str(yr)[-2:]}")
            c.font = HEADER_FONT; c.fill = HEADER_FILL
            c.alignment = Alignment(horizontal="center"); c.border = BORDER
        row += 1
        amt_start = row
        for idx, e in enumerate(emis, 1):
            ws.cell(row=row, column=1, value=idx).alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=1).border = BORDER
            ws.cell(row=row, column=2, value=e.lender).font = NORMAL
            ws.cell(row=row, column=2).border = BORDER
            by_ym = {(it["date"].year, it["date"].month): it["amount"] for it in e.items}
            for col_idx, (yr, mo) in enumerate(months, start=3):
                c = ws.cell(row=row, column=col_idx)
                if (yr, mo) in by_ym:
                    c.value = by_ym[(yr, mo)]
                    c.number_format = '#,##0.00'
                    c.alignment = Alignment(horizontal="right"); c.font = NORMAL
                c.border = BORDER
            row += 1

        # Monthly total
        ws.cell(row=row, column=2, value="Monthly Total EMI (₹)").font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
        ws.cell(row=row, column=2).fill = HEADER_FILL
        ws.cell(row=row, column=1).fill = HEADER_FILL
        ws.cell(row=row, column=1).border = BORDER
        ws.cell(row=row, column=2).border = BORDER
        for col_idx in range(3, 3 + len(months)):
            col_letter = get_column_letter(col_idx)
            c = ws.cell(row=row, column=col_idx, value=f"=SUM({col_letter}{amt_start}:{col_letter}{amt_start + len(emis) - 1})")
            c.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
            c.fill = HEADER_FILL; c.number_format = '#,##0.00'
            c.alignment = Alignment(horizontal="right"); c.border = BORDER

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 36
    for col_l in ["C", "D"]:
        ws.column_dimensions[col_l].width = 14
    for col_l in ["E", "F", "G"]:
        ws.column_dimensions[col_l].width = 11
    for col_l in ["H", "I"]:
        ws.column_dimensions[col_l].width = 12
    for col_l in ["J", "K", "L"]:
        ws.column_dimensions[col_l].width = 14
    ws.column_dimensions["M"].width = 12
    ws.column_dimensions["N"].width = 16
    ws.column_dimensions["O"].width = 40
    # NEW: loan estimation cols
    ws.column_dimensions["P"].width = 28   # Loan Type
    ws.column_dimensions["Q"].width = 11   # Confidence
    ws.column_dimensions["R"].width = 17   # Principal Low
    ws.column_dimensions["S"].width = 17   # Principal Mid
    ws.column_dimensions["T"].width = 17   # Principal High

    # Assumptions footer
    if emis:
        from analyzers.loan_estimate import LOAN_PROFILES
        row += 2
        note_cell = ws.cell(row=row, column=1,
                            value="Assumptions used for loan principal estimation:")
        note_cell.font = SUBTITLE
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=20)
        row += 1
        for key, p in LOAN_PROFILES.items():
            note = (f"  • {p['label']}: assumed rate {p['rate_low']*100:.1f}-{p['rate_high']*100:.1f}% p.a., "
                    f"tenure {p['tenure_low']}-{p['tenure_high']} months")
            ws.cell(row=row, column=1, value=note).font = Font(name=FONT, size=9, color="595959")
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=20)
            row += 1
        disc = ws.cell(row=row, column=1,
                       value=("Disclaimers: (1) These are ORIGINAL principal estimates - current outstanding is lower due to amortization. "
                              "(2) Tenure and rate are inferred from loan-type heuristics, not statement data. "
                              "(3) Floating-rate or restructured loans may deviate from estimates. "
                              "(4) Use Low/High range to bound risk; Mid is the most-likely point estimate."))
        disc.font = Font(name=FONT, italic=True, size=9, color="595959")
        disc.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=20)
        ws.row_dimensions[row].height = 50


def _build_summary_sheet(wb, stmt: ParsedStatement, abb: ABBResult, emis: List[DetectedEMI]):
    ws = wb.create_sheet("Summary")
    M = stmt.metadata
    last_col_letter = get_column_letter(len(abb.months) + 1)

    ws["A1"] = "Bank Statement Analysis - Summary"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:B1")

    row = 3
    ws.cell(row=row, column=1, value="ACCOUNT INFORMATION").font = HEADER_FONT
    ws.cell(row=row, column=1).fill = HEADER_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    row += 1

    info = [
        ("Bank", M.bank),
        ("Account Holder", M.account_holder),
        ("Account No.", M.account_no),
        ("Account Type", M.account_type),
        ("Statement Period",
         f"{M.period_from.strftime('%d-%b-%Y') if M.period_from else ''} to {M.period_to.strftime('%d-%b-%Y') if M.period_to else ''}"),
        ("Total Days", len(abb.daily_eod)),
        ("Days with Transactions", len(abb.daily_eod) - abb.days_carry_fwd),
        ("Days Carry-Forward", abb.days_carry_fwd),
        ("Opening Balance ₹", M.opening_balance),
        ("Closing Balance ₹", M.closing_balance),
        ("Total Debit ₹", M.total_debit),
        ("Total Credit ₹", M.total_credit),
    ]
    for label, val in info:
        ws.cell(row=row, column=1, value=label).font = BOLD
        ws.cell(row=row, column=1).fill = SUB_FILL
        c = ws.cell(row=row, column=2, value=val); c.font = NORMAL
        if isinstance(val, float):
            c.number_format = '"₹"#,##0.00'
        ws.cell(row=row, column=1).border = BORDER
        ws.cell(row=row, column=2).border = BORDER
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="ABB METRICS").font = HEADER_FONT
    ws.cell(row=row, column=1).fill = HEADER_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    row += 1

    matrix_range = f"'Daily Balance Matrix'!B5:{last_col_letter}35"
    metrics = [
        ("Total Days Counted", f'=COUNT({matrix_range})', '#,##0'),
        ("Average Bank Balance (ABB) ₹", f'=AVERAGE({matrix_range})', '"₹"#,##0.00'),
        ("Median EOD ₹", f'=MEDIAN({matrix_range})', '"₹"#,##0.00'),
        ("Minimum EOD ₹", f'=MIN({matrix_range})', '"₹"#,##0.00'),
        ("Maximum EOD ₹", f'=MAX({matrix_range})', '"₹"#,##0.00'),
        ("Days with Balance < ₹1,000", f'=COUNTIF({matrix_range},"<1000")', '#,##0'),
        ("Days with Balance < ₹500", f'=COUNTIF({matrix_range},"<500")', '#,##0'),
        ("Days with Balance ≥ ₹50,000", f'=COUNTIF({matrix_range},">=50000")', '#,##0'),
        ("Days with Balance ≥ ₹1,00,000", f'=COUNTIF({matrix_range},">=100000")', '#,##0'),
    ]
    for label, formula, fmt in metrics:
        ws.cell(row=row, column=1, value=label).font = BOLD
        ws.cell(row=row, column=1).fill = SUB_FILL
        c = ws.cell(row=row, column=2, value=formula); c.font = NORMAL; c.number_format = fmt
        ws.cell(row=row, column=1).border = BORDER
        ws.cell(row=row, column=2).border = BORDER
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="EMI OBLIGATION SUMMARY").font = HEADER_FONT
    ws.cell(row=row, column=1).fill = HEADER_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    row += 1
    ws.cell(row=row, column=1, value="Number of EMIs Detected").font = BOLD
    ws.cell(row=row, column=1).fill = SUB_FILL
    ws.cell(row=row, column=2, value=len(emis)).number_format = '#,##0'
    ws.cell(row=row, column=1).border = BORDER
    ws.cell(row=row, column=2).border = BORDER
    row += 1
    if emis:
        ws.cell(row=row, column=1, value="Total Monthly EMI (sum of avg) ₹").font = BOLD
        ws.cell(row=row, column=1).fill = SUB_FILL
        emi_start = 6  # in EMI Extract: row 6 is first EMI (header row 5)
        emi_end = emi_start + len(emis) - 1
        c = ws.cell(row=row, column=2, value=f"=SUM('EMI Extract'!J{emi_start}:J{emi_end})")
        c.number_format = '"₹"#,##0.00'
        ws.cell(row=row, column=1).border = BORDER
        ws.cell(row=row, column=2).border = BORDER
        row += 1

        # NEW: Estimated original principal range
        from analyzers.loan_estimate import total_exposure
        exp = total_exposure(emis)

        ws.cell(row=row, column=1, value="True Monthly Obligation ₹").font = BOLD
        ws.cell(row=row, column=1).fill = SUB_FILL
        c = ws.cell(row=row, column=2, value=exp["total_monthly_emi"])
        c.number_format = '"₹"#,##0.00'; c.font = BOLD
        ws.cell(row=row, column=1).border = BORDER
        ws.cell(row=row, column=2).border = BORDER
        row += 2

        ws.cell(row=row, column=1, value="ESTIMATED ORIGINAL DEBT EXPOSURE").font = HEADER_FONT
        ws.cell(row=row, column=1).fill = HEADER_FILL
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        row += 1

        for label, val, fmt, weight in [
            ("Principal (Low estimate) ₹", exp["total_principal_low"], '"₹"#,##0', "normal"),
            ("Principal (Mid estimate) ₹  ← most likely", exp["total_principal_mid"], '"₹"#,##0', "bold"),
            ("Principal (High estimate) ₹", exp["total_principal_high"], '"₹"#,##0', "normal"),
        ]:
            ws.cell(row=row, column=1, value=label).font = BOLD
            ws.cell(row=row, column=1).fill = SUB_FILL
            c = ws.cell(row=row, column=2, value=val)
            c.number_format = fmt
            c.font = BOLD if weight == "bold" else NORMAL
            if weight == "bold":
                c.fill = PatternFill("solid", start_color="FFEB9C")
            ws.cell(row=row, column=1).border = BORDER
            ws.cell(row=row, column=2).border = BORDER
            row += 1
        # Note
        row += 1
        note_cell = ws.cell(row=row, column=1,
                            value=("Note: ORIGINAL principal estimates; current outstanding will be lower. "
                                   "Tenure/rate inferred from loan-type heuristics. See EMI Extract sheet for per-loan detail."))
        note_cell.font = Font(name=FONT, italic=True, size=9, color="595959")
        note_cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        ws.row_dimensions[row].height = 40

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 28


def _build_chart_sheet(wb, abb: ABBResult):
    ws = wb.create_sheet("Balance Chart")
    ws["A1"] = "Daily EOD Balance Trend"
    ws["A1"].font = TITLE_FONT
    ws.cell(row=3, column=1, value="Date").font = HEADER_FONT
    ws.cell(row=3, column=1).fill = HEADER_FILL
    ws.cell(row=3, column=2, value="EOD Balance").font = HEADER_FONT
    ws.cell(row=3, column=2).fill = HEADER_FILL

    sorted_dates = sorted(abb.daily_eod.keys())
    for i, d in enumerate(sorted_dates, start=4):
        ws.cell(row=i, column=1, value=d).number_format = "dd-mmm-yy"
        ws.cell(row=i, column=2, value=abb.daily_eod[d]).number_format = '#,##0.00'

    n = len(sorted_dates)
    chart = LineChart()
    chart.title = f"Daily EOD Balance ({abb.period_start.strftime('%d-%b-%Y')} to {abb.period_end.strftime('%d-%b-%Y')})"
    chart.style = 2
    chart.y_axis.title = "Balance (₹)"; chart.x_axis.title = "Date"
    chart.height = 15; chart.width = 32; chart.legend = None
    data_ref = Reference(ws, min_col=2, min_row=3, max_row=n + 3, max_col=2)
    cat_ref = Reference(ws, min_col=1, min_row=4, max_row=n + 3)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cat_ref)
    if chart.series:
        chart.series[0].graphicalProperties.line.solidFill = "1F4E78"
        chart.series[0].graphicalProperties.line.width = 12700
    ws.add_chart(chart, "D3")
    ws.column_dimensions["A"].width = 13
    ws.column_dimensions["B"].width = 16


def _build_party_sheet(wb, credit_parties: List[PartyAggregate],
                       debit_parties: List[PartyAggregate], months):
    """Party-wise analysis: top counterparties + month-grid for credits & debits."""
    ws = wb.create_sheet("Party-wise")
    n_months = len(months)

    ws["A1"] = "PARTY-WISE ANALYSIS"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_months + 5)

    ws["A2"] = ("Top counterparties identified from transaction narrations (UPI / NEFT / IMPS / RTGS / TPT / ECS / ACH). "
                "Type column classifies for credit-underwriting use. Parties with <2 transactions AND <₹10K total are rolled into 'Other'.")
    ws["A2"].font = Font(name=FONT, italic=True, size=9, color="595959")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_months + 5)

    row = 4

    # ============ SECTION A: CREDIT (Receivers) ============
    row = _render_party_section(
        ws, row, credit_parties, months,
        title="Section A: Top RECEIVERS (Credit Parties — Money Received INTO Account)",
        amount_label="Total Received ₹",
        section_fill=CREDIT_FILL,
    )

    row += 2

    # ============ SECTION B: DEBIT (Payers) ============
    row = _render_party_section(
        ws, row, debit_parties, months,
        title="Section B: Top PAYERS (Debit Parties — Money Sent FROM Account)",
        amount_label="Total Paid ₹",
        section_fill=DEBIT_FILL,
    )

    # Column widths
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 22   # Type column
    ws.column_dimensions["E"].width = 8    # Txns
    ws.column_dimensions["F"].width = 8    # Months
    ws.column_dimensions["G"].width = 14   # Total
    ws.column_dimensions["H"].width = 12   # Avg
    ws.column_dimensions["I"].width = 14   # Largest
    ws.column_dimensions["J"].width = 12   # First Date
    ws.column_dimensions["K"].width = 12   # Last Date
    for i in range(12, 12 + n_months):
        ws.column_dimensions[get_column_letter(i)].width = 13


def _render_party_section(ws, start_row, parties, months, title, amount_label, section_fill):
    """Render one section (Credit OR Debit). Returns next available row.
    Columns: # | Party | Mode | Type | Txns | Months | Total ₹ | Avg/Txn | Largest | First | Last | <month grid>
    """
    n_months = len(months)
    row = start_row

    # Section title
    tc = ws.cell(row=row, column=1, value=title)
    tc.font = SUBTITLE
    tc.fill = section_fill
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=11 + n_months)
    row += 1

    # Subsection: Summary Table
    summary_headers = [
        "#", "Party", "Mode", "Type", "Txns", "Months",
        amount_label, "Avg/Txn ₹", "Largest ₹", "First Date", "Last Date",
    ]
    for i, h in enumerate(summary_headers):
        c = ws.cell(row=row, column=i + 1, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    # Month-grid headers
    for col_idx, (yr, mo) in enumerate(months, start=12):
        c = ws.cell(row=row, column=col_idx, value=f"{MONTH_NAMES[mo]}-{str(yr)[-2:]}")
        c.font = HEADER_FONT; c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER
    ws.row_dimensions[row].height = 30
    row += 1

    if not parties:
        ws.cell(row=row, column=1, value="No parties detected in this category.").font = Font(name=FONT, italic=True, size=10, color="595959")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=11 + n_months)
        return row + 1

    # Color coding for Type column
    TYPE_FILLS = {
        "Business": PatternFill("solid", start_color="C6EFCE"),
        "Related Party / Self": PatternFill("solid", start_color="FFEB9C"),
        "EMI / Loan": PatternFill("solid", start_color="FFC7CE"),
        "Bank Charges": PatternFill("solid", start_color="DDDDDD"),
        "Tax / Statutory": PatternFill("solid", start_color="D9E1F2"),
        "Cash": PatternFill("solid", start_color="FCE4D6"),
        "Salary / Wages": PatternFill("solid", start_color="E2EFDA"),
        "Cheque Deposit": PatternFill("solid", start_color="DDEBF7"),
        "Cheque Return": PatternFill("solid", start_color="F8CBAD"),
        "Vehicle Expense (FASTag)": PatternFill("solid", start_color="FFF2CC"),
    }

    data_start_row = row
    for idx, p in enumerate(parties, start=1):
        ws.cell(row=row, column=1, value=idx).alignment = Alignment(horizontal="center")
        pc = ws.cell(row=row, column=2, value=p.party)
        pc.font = NORMAL
        if p.party.startswith("Other"):
            pc.font = Font(name=FONT, italic=True, size=10, color="595959")
        ws.cell(row=row, column=3, value=p.mode).alignment = Alignment(horizontal="center")

        # Type column with color coding
        tc_cell = ws.cell(row=row, column=4, value=p.txn_type)
        tc_cell.alignment = Alignment(horizontal="center")
        tc_cell.font = Font(name=FONT, size=9)
        if p.txn_type in TYPE_FILLS:
            tc_cell.fill = TYPE_FILLS[p.txn_type]

        ws.cell(row=row, column=5, value=p.txn_count).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=6, value=p.months_active).alignment = Alignment(horizontal="center")

        tc = ws.cell(row=row, column=7, value=p.total_amount)
        tc.number_format = '#,##0.00'
        tc.alignment = Alignment(horizontal="right")
        tc.font = BOLD

        ac = ws.cell(row=row, column=8,
                     value=(p.total_amount / p.txn_count if p.txn_count else 0))
        ac.number_format = '#,##0.00'; ac.alignment = Alignment(horizontal="right")

        lc = ws.cell(row=row, column=9, value=p.largest_txn)
        lc.number_format = '#,##0.00'; lc.alignment = Alignment(horizontal="right")

        if p.first_date:
            fc = ws.cell(row=row, column=10, value=p.first_date)
            fc.number_format = "dd-mmm-yy"
            fc.alignment = Alignment(horizontal="center")
        if p.last_date:
            ld = ws.cell(row=row, column=11, value=p.last_date)
            ld.number_format = "dd-mmm-yy"
            ld.alignment = Alignment(horizontal="center")

        # Month-by-month amounts (start at col 12)
        for col_idx, (yr, mo) in enumerate(months, start=12):
            amt = p.by_month.get((yr, mo))
            if amt and amt > 0:
                c = ws.cell(row=row, column=col_idx, value=amt)
                c.number_format = '#,##0.00'
                c.alignment = Alignment(horizontal="right")
                c.font = NORMAL
                c.fill = section_fill
            ws.cell(row=row, column=col_idx).border = BORDER

        for col in range(1, 12):
            ws.cell(row=row, column=col).border = BORDER
            if ws.cell(row=row, column=col).font.size is None:
                ws.cell(row=row, column=col).font = NORMAL
        row += 1

    # Totals row (formulas)
    data_end_row = row - 1
    ws.cell(row=row, column=2, value="TOTAL").font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
    ws.cell(row=row, column=2).fill = HEADER_FILL
    for col in [1, 3, 4, 6, 8, 9, 10, 11]:
        ws.cell(row=row, column=col).fill = HEADER_FILL
        ws.cell(row=row, column=col).border = BORDER
    # Count txns total (col 5 now)
    c = ws.cell(row=row, column=5, value=f"=SUM(E{data_start_row}:E{data_end_row})")
    c.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
    c.fill = HEADER_FILL; c.number_format = '#,##0'
    c.alignment = Alignment(horizontal="center"); c.border = BORDER
    # Total amount column (col 7 now)
    tc = ws.cell(row=row, column=7, value=f"=SUM(G{data_start_row}:G{data_end_row})")
    tc.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
    tc.fill = HEADER_FILL; tc.number_format = '"₹"#,##0.00'
    tc.alignment = Alignment(horizontal="right"); tc.border = BORDER
    # Per-month totals (col 12+)
    for col_idx in range(12, 12 + n_months):
        col_letter = get_column_letter(col_idx)
        c = ws.cell(row=row, column=col_idx,
                    value=f"=SUM({col_letter}{data_start_row}:{col_letter}{data_end_row})")
        c.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
        c.fill = HEADER_FILL; c.number_format = '#,##0.00'
        c.alignment = Alignment(horizontal="right"); c.border = BORDER
    return row + 1


def _build_monthly_summary_sheet(wb, monthly_stats):
    """Per-month summary: Open, Close, Cr/Dr amounts & counts, charges, returns, ABB."""
    ws = wb.create_sheet("Monthly Summary")

    ws["A1"] = "MONTHLY CASH FLOW SUMMARY"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11)

    ws["A2"] = ("Per-month aggregation of credits, debits, bank charges, and bounce events. "
                "Used by underwriting teams to assess monthly business cash flow.")
    ws["A2"].font = Font(name=FONT, italic=True, size=9, color="595959")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=11)

    headers = [
        "Month", "Open Bal ₹", "Close Bal ₹",
        "Total Credit ₹", "# Cr Txns",
        "Total Debit ₹", "# Dr Txns",
        "Bank Charges ₹",
        "Inward Returns #", "Outward Returns #", "ECS/NACH Returns #",
    ]
    row = 4
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = HEADER_FONT; c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[row].height = 30
    row += 1

    data_start = row
    for m in monthly_stats:
        ws.cell(row=row, column=1, value=f"{MONTH_NAMES[m.month]}-{str(m.year)[-2:]}").alignment = Alignment(horizontal="center")
        for col, val, fmt in [
            (2, m.opening_balance, '#,##0.00'),
            (3, m.closing_balance, '#,##0.00'),
            (4, m.cr_amount, '#,##0.00'),
            (5, m.cr_count, '#,##0'),
            (6, m.dr_amount, '#,##0.00'),
            (7, m.dr_count, '#,##0'),
            (8, m.bank_charges_amount, '#,##0.00'),
            (9, m.inward_returns_count, '#,##0'),
            (10, m.outward_returns_count, '#,##0'),
            (11, m.ecs_nach_returns_count, '#,##0'),
        ]:
            c = ws.cell(row=row, column=col, value=val if val is not None else 0)
            c.number_format = fmt
            c.alignment = Alignment(horizontal="right" if col >= 2 else "center")
            c.font = NORMAL
            c.border = BORDER
        # Color-code returns columns when > 0
        for col in [9, 10, 11]:
            v = ws.cell(row=row, column=col).value or 0
            if v > 0:
                ws.cell(row=row, column=col).fill = PatternFill("solid", start_color="FFC7CE")
                ws.cell(row=row, column=col).font = BOLD
        # Color charges if high
        ws.cell(row=row, column=1).border = BORDER
        ws.cell(row=row, column=1).font = BOLD
        row += 1

    data_end = row - 1

    # TOTAL row using SUM formulas
    tc = ws.cell(row=row, column=1, value="TOTAL")
    tc.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
    tc.fill = HEADER_FILL; tc.alignment = Alignment(horizontal="center"); tc.border = BORDER

    for col in [2, 3]:  # Open/Close: don't sum these
        ws.cell(row=row, column=col).fill = HEADER_FILL
        ws.cell(row=row, column=col).border = BORDER
    for col, fmt in [
        (4, '#,##0.00'), (5, '#,##0'), (6, '#,##0.00'), (7, '#,##0'),
        (8, '#,##0.00'), (9, '#,##0'), (10, '#,##0'), (11, '#,##0'),
    ]:
        col_letter = get_column_letter(col)
        c = ws.cell(row=row, column=col, value=f"=SUM({col_letter}{data_start}:{col_letter}{data_end})")
        c.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
        c.fill = HEADER_FILL; c.number_format = fmt
        c.alignment = Alignment(horizontal="right"); c.border = BORDER

    # Column widths
    ws.column_dimensions["A"].width = 12
    for col in range(2, 12):
        ws.column_dimensions[get_column_letter(col)].width = 16
    ws.freeze_panes = "B5"


def _build_returns_sheet(wb, returns_data):
    """Listing of all cheque/ECS/NACH returns + summary block."""
    ws = wb.create_sheet("Cheque Returns")

    # Title
    ws["A1"] = "CHEQUE / ECS / NACH RETURNS — BOUNCE TRACK RECORD"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

    ws["A2"] = ("Inward = cheque you deposited bounced. Outward = cheque you issued bounced. "
                "ECS/NACH return = EMI mandate failure (critical underwriting flag).")
    ws["A2"].font = Font(name=FONT, italic=True, size=9, color="595959")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)

    # Summary block at top
    from analyzers.returns import returns_summary
    summary = returns_summary(returns_data)

    row = 4
    ws.cell(row=row, column=1, value="SUMMARY").font = SUBTITLE
    ws.cell(row=row, column=1).fill = HEADER_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1

    summary_rows = [
        ("Inward Cheque Returns (received cheques that bounced)", summary["inward_count"], summary["inward_amount"]),
        ("Outward Cheque Returns (your issued cheques that bounced)", summary["outward_count"], summary["outward_amount"]),
        ("ECS / NACH / ACH Returns (mandate / EMI failures)", summary["ecs_nach_count"], summary["ecs_nach_amount"]),
        ("Return-related Charges", summary["charge_count"], summary["total_charges"]),
    ]
    ws.cell(row=row, column=1, value="Category").font = HEADER_FONT
    ws.cell(row=row, column=1).fill = HEADER_FILL; ws.cell(row=row, column=1).border = BORDER
    ws.cell(row=row, column=4, value="Count").font = HEADER_FONT
    ws.cell(row=row, column=4).fill = HEADER_FILL; ws.cell(row=row, column=4).border = BORDER
    ws.cell(row=row, column=5, value="Amount ₹").font = HEADER_FONT
    ws.cell(row=row, column=5).fill = HEADER_FILL; ws.cell(row=row, column=5).border = BORDER
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    row += 1
    for label, count, amt in summary_rows:
        lc = ws.cell(row=row, column=1, value=label)
        lc.font = NORMAL; lc.border = BORDER
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        cc = ws.cell(row=row, column=4, value=count)
        cc.number_format = '#,##0'; cc.alignment = Alignment(horizontal="center")
        cc.font = BOLD; cc.border = BORDER
        if count > 0 and "Charges" not in label:
            cc.fill = PatternFill("solid", start_color="FFC7CE")
        ac = ws.cell(row=row, column=5, value=amt)
        ac.number_format = '#,##0.00'; ac.alignment = Alignment(horizontal="right")
        ac.font = BOLD; ac.border = BORDER
        row += 1

    row += 2

    # Detail listing
    if not returns_data:
        ws.cell(row=row, column=1,
                value="✓ NO BOUNCES DETECTED — clean track record across the statement period.").font = Font(
            name=FONT, bold=True, size=11, color="008000")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    else:
        ws.cell(row=row, column=1, value="DETAIL LISTING").font = SUBTITLE
        ws.cell(row=row, column=1).fill = HEADER_FILL
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        row += 1

        headers = ["#", "Date", "Type", "Amount ₹", "Cheque No", "Narration"]
        for i, h in enumerate(headers, start=1):
            c = ws.cell(row=row, column=i, value=h)
            c.font = HEADER_FONT; c.fill = HEADER_FILL
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = BORDER
        row += 1

        for idx, r in enumerate(returns_data, start=1):
            ws.cell(row=row, column=1, value=idx).alignment = Alignment(horizontal="center")
            dc = ws.cell(row=row, column=2, value=r.date)
            dc.number_format = "dd-mmm-yy"; dc.alignment = Alignment(horizontal="center")
            type_label = r.return_type + (" (charge)" if r.is_charge else "")
            tc = ws.cell(row=row, column=3, value=type_label)
            tc.alignment = Alignment(horizontal="center")
            if not r.is_charge:
                tc.fill = PatternFill("solid", start_color="FFC7CE")
            ac = ws.cell(row=row, column=4, value=r.amount)
            ac.number_format = '#,##0.00'; ac.alignment = Alignment(horizontal="right")
            ac.font = BOLD
            cc = ws.cell(row=row, column=5, value=r.cheque_no or "—")
            cc.alignment = Alignment(horizontal="center")
            nc = ws.cell(row=row, column=6, value=r.narration)
            nc.font = Font(name=FONT, size=9)
            nc.alignment = Alignment(horizontal="left", wrap_text=True)
            for col in range(1, 7):
                ws.cell(row=row, column=col).border = BORDER
            row += 1

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 70
